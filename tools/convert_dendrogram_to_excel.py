#!/usr/bin/env python3
"""
Convert dendrogram_categories.json to a hierarchical Excel table with full hierarchy
"""

import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def convert_dendrogram_to_excel(json_path, excel_path):
    """Convert dendrogram categories JSON to hierarchical Excel table"""

    # Read JSON file
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Define the full hierarchy mapping based on drawing.png
    hierarchy = {
        'People': {
            'level': 1,
            'subcategories': {
                'Integrated Systems Design': [
                    'Team Collaboration & Feedback Systems',
                    'Cross-Functional Organizational Design',
                    'Team Empowerment & Alignment',
                    'Adaptive Architecture & Planning',
                    'Design for Integration & Simplicity',
                    'Systems Engineering & Integration'
                ],
                'Operational Excellence': [
                    'Lean Culture & Quality',
                    'Project Management & Stakeholder Visibility',
                    'Decision Analytics & Strategy',
                    'Flow, Process Optimization & Capacity',
                    'Leadership Development & Coaching'
                ]
            }
        },
        'Process': {
            'level': 1,
            'subcategories': {
                'Operational Excellence': [
                    'Continuous Improvement Cycles',
                    'Real-Time Adaptive Problem Solving'
                ],
                'Organizational Agility': [
                    'Agile Scaling & Transformation',
                    'Continuous Delivery & Experimentation',
                    'Customer-Centric Incremental Delivery'
                ]
            }
        }
    }

    # Create a mapping of categories to their Level 2 and Level 1 parents
    category_to_parents = {}
    for l1_name, l1_data in hierarchy.items():
        for l2_name, l3_categories in l1_data['subcategories'].items():
            for l3_name in l3_categories:
                category_to_parents[l3_name] = {
                    'level1': l1_name,
                    'level2': l2_name
                }

    # Prepare data for Excel
    rows = []

    # Group categories by Level 1 and Level 2
    for l1_name, l1_data in hierarchy.items():
        # Add Level 1 row
        rows.append({
            'Level 1': l1_name,
            'Level 2': '',
            'Level 3 (Category)': '',
            'Level 4 (Cluster)': '',
            'Strength': '',
            'Bonus': '',
            'Cluster_ID': '',
            'Cluster_Count': ''
        })

        for l2_name, l3_categories in l1_data['subcategories'].items():
            # Add Level 2 row
            rows.append({
                'Level 1': '',
                'Level 2': l2_name,
                'Level 3 (Category)': '',
                'Level 4 (Cluster)': '',
                'Strength': '',
                'Bonus': '',
                'Cluster_ID': '',
                'Cluster_Count': ''
            })

            # Add Level 3 (categories) and Level 4 (clusters)
            for category in data['categories']:
                if category['name'] in l3_categories:
                    # Add category row
                    rows.append({
                        'Level 1': '',
                        'Level 2': '',
                        'Level 3 (Category)': category['name'],
                        'Level 4 (Cluster)': '',
                        'Strength': category['strength'],
                        'Bonus': category['bonus'],
                        'Cluster_ID': '',
                        'Cluster_Count': len(category['clusters'])
                    })

                    # Add cluster rows
                    for cluster_id, cluster_name in zip(category['clusters'], category['cluster_names']):
                        rows.append({
                            'Level 1': '',
                            'Level 2': '',
                            'Level 3 (Category)': '',
                            'Level 4 (Cluster)': cluster_name,
                            'Strength': '',
                            'Bonus': '',
                            'Cluster_ID': cluster_id,
                            'Cluster_Count': ''
                        })

    # Create DataFrame
    df = pd.DataFrame(rows)

    # Write to Excel
    df.to_excel(excel_path, index=False, sheet_name='Hierarchy')

    # Format the Excel file
    wb = load_workbook(excel_path)
    ws = wb['Hierarchy']

    # Define styles (matching the drawing.png colors)
    level1_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')  # Blue
    level2_fill = PatternFill(start_color='FF6DB6', end_color='FF6DB6', fill_type='solid')  # Pink/Magenta
    level3_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
    level4_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Light Gray
    header_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')

    level1_font = Font(color='FFFFFF', bold=True, size=14)
    level2_font = Font(color='FFFFFF', bold=True, size=12)
    level3_font = Font(color='FFFFFF', bold=True, size=11)
    level4_font = Font(color='000000', size=10)
    header_font = Font(color='FFFFFF', bold=True, size=11)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Format data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Determine which level this row is
        level1_val = row[0].value
        level2_val = row[1].value
        level3_val = row[2].value
        level4_val = row[3].value

        if level1_val:  # Level 1 row
            for cell in row:
                cell.fill = level1_fill
                cell.font = level1_font
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = border
        elif level2_val:  # Level 2 row
            for cell in row:
                cell.fill = level2_fill
                cell.font = level2_font
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
                cell.border = border
        elif level3_val:  # Level 3 (Category) row
            for cell in row:
                cell.fill = level3_fill
                cell.font = level3_font
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)
                cell.border = border
        elif level4_val:  # Level 4 (Cluster) row
            for cell in row:
                cell.fill = level4_fill
                cell.font = level4_font
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=3)
                cell.border = border

    # Adjust column widths
    ws.column_dimensions['A'].width = 25  # Level 1
    ws.column_dimensions['B'].width = 35  # Level 2
    ws.column_dimensions['C'].width = 45  # Level 3 (Category)
    ws.column_dimensions['D'].width = 60  # Level 4 (Cluster)
    ws.column_dimensions['E'].width = 12  # Strength
    ws.column_dimensions['F'].width = 10  # Bonus
    ws.column_dimensions['G'].width = 12  # Cluster_ID
    ws.column_dimensions['H'].width = 15  # Cluster_Count

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save workbook
    wb.save(excel_path)
    print(f"✓ Excel file created: {excel_path}")
    print(f"  Total categories: {len(data['categories'])}")
    print(f"  Total rows: {len(rows)}")


if __name__ == '__main__':
    json_path = 'results_semantic_clustering_combined/dendrogram_categories.json'
    excel_path = 'results_semantic_clustering_combined/dendrogram_categories.xlsx'

    convert_dendrogram_to_excel(json_path, excel_path)
